# translate a SIX FLEX dictionary to EDM Saas DD

# to work with excel
from openpyxl import load_workbook 
# to work with dataframes
import pandas as pd
# to show logs
from loguru import logger
# to show progress
from tqdm import tqdm


def main():
    
    # file name and path location data
    six_path = 'C:/Users/martin.thomasmathew/OneDrive - S&P Global/Documents/Python Scripts/SIX/'
    six_filename = 'SIX-Flex-package-description.xlsx'

    out_file_path = 'C:/Users/martin.thomasmathew/OneDrive - S&P Global/Documents/Python Scripts/SIX/'
    out_file_name = 'six-flex_EDMSaasDD.csv'

    # global variable - column in any sheet where word: 'Pos' is likely to occur
    global pos_col
    pos_col = 2

    # load excel into into object
    workbook = load_workbook(six_path + six_filename)

    # get all sheets with 'Pos' in 2nd column
    sheets = get_pos_sheets(workbook, 'Pos')

    # aggregate dataframe from all sheets
    agg_df = process_sheets(workbook,sheets)

    # convert excel df to EDM Saas DF
    saas_dd_df = convert_df_to_saas_format(agg_df)

    # save df as csv
    saas_dd_df.to_csv(out_file_path + out_file_name, index=False)








    
    
##################################################################################################################
##################################################################################################################
##################################################################################################################
##################################################################################################################

def convert_df_to_saas_format(df):
    '''
    Args:
    df - dataframe of content from excel

    Returns
    dataframe that can be imported to EDM Saas DD    
    '''

    # initialize the Saas DD df
    saas_dd_col_list = ['Dictionary Name', 'Dictionary Description', 'Entity Name', 'Entity Description', 'Attribute Name'
                    , 'Attribute Description', 'Attribute DataType', 'Attribute Precision', 'Attribute Scale'
                    , 'Attribute Max Length', 'Attribute Field Name', 'Attribute Short Description'
                    , 'Attribute Technical Type', 'Attribute Is Key', 'Attribute Required'
                    ,  'Attribute Is Nullable', 'Attribute Sensitive Data', 'Attribute Default Value'
                    , 'Business Data Type',]
    
    saas_df = pd.DataFrame(columns = saas_dd_col_list)

    # mapping the simple items
    smld = ['Entity', 'Entity', 'Field Name', 'Field Description', 'CSV Header', 'Field Description' ] # smld - simple map items list from six flex df
    smls = ['Entity Name', 'Entity Description','Attribute Name', 'Attribute Description', 'Attribute Field Name'
         , 'Attribute Short Description'] # smls - simple map items list for Saas DD df
    
    for count,s in enumerate(smls):
        saas_df[s] = df[smld[count]]
    
    # default values from six flex df to saas df
    saas_df['Dictionary Name'] = 'SIX'
    saas_df['Dictionary Description'] = 'SIX FLEX'
    saas_df['Attribute Sensitive Data'] = 0

    # tranform values of six flex df of column field type to saas df column - attribute data type
    saas_df['Attribute DataType']= df['Field Type'].apply(transform_value)

    # transform values of six flex df of column - Result Key(s) to saas df column - attribute is key
    saas_df['Attribute Is Key'] = df['Result Key(s)'].apply(transform_value_2)

    # copy values of 1 column to another column in same df 
    saas_df['Attribute Technical Type'] = saas_df['Attribute DataType']
    saas_df['Attribute Required'] = saas_df['Attribute Is Key']

    # inverse value in 1 column based on value in another column in saas df
    saas_df['Attribute Is Nullable']= 1- saas_df['Attribute Is Key']

    return saas_df




def transform_value_2 (value):
    '''
    value - value of cell from excel 
    return transformed value
    '''
    if value is None:
        return 0
    elif value == 'Key':
        return 1




def transform_value(value):
    '''
    value - cell value from column in excel
    
    Returns:
    returns transformed value per rules
    '''
    
    if value in ['Integer', 'Date']:
        return value
    elif value == None:
        return 'String'
    elif value.startswith('Enumeration'):
        return 'String'
    elif value == 'Real':
        return 'Decimal'
    elif value == 'Day of a Specific Month':
        return 'Integer'
    elif value in ('Time Of Day Exact To The Second', 'Time Of Day Exact To The Minute', 'Date And Time Exact To The Second' ):
        return 'Datetime'
    else:
        return value


def process_sheets(wb, sheetlist):
    '''
    Arg:
    wb - workbook obj of excel
    sheetlist - list of all sheets with Pos in them
    Returns:
    a dataframe that concatenates all the dataframes from all the input sheets
    '''

    wb_list = [] # wb_list - list. Items are dataframes. 1 per sheet

    sheet_count = len(sheetlist)


    for sheet in tqdm(sheetlist, desc='Sheet processing', total=sheet_count):
        dic_sheet = getrowrange(wb, sheet)
        print(sheet,'\n', dic_sheet) # just for test. best to have logger here
        sheet_df = getsheetdata(wb,sheet, dic_sheet)
        logger.info(f'Sheet:{sheet} is processed')
        wb_list.append(sheet_df)
    

    excel_df = pd.concat(wb_list, ignore_index=True)

    return excel_df
               
    



def getsheetdata(wb,s,d):
    '''
    Arg:
    wb - workbook
    s - sheet
    d - dictionary of entity as key and value is list with 2 items - row start and row end for a grid in the sheet

    Returns:
    dataframe that concatenates all the dataframes of the tables from a single sheet
    '''

    s_df_list = [] # s_df_list - list for all the dataframes for a sheet

    sheet_obj = wb[s]

    field_length_col = 9 # field_length_col is the column number with the last column of grid in excel

    for k in d:
        row_range = d[k]
        rows = []
        for row in sheet_obj.iter_rows(min_row = row_range[0], max_row = row_range[1]
                                       ,min_col = pos_col, max_col = field_length_col, values_only = True ):
            rows.append(row)
        
        # create dataframe of the rows from excel sheet    
        df = pd.DataFrame(rows[1:], columns = rows[0])
        
        # add entity column and dic key
        df.insert(0, 'Entity', k, True)

        s_df_list.append(df)
    
    s_df = pd.concat(s_df_list, ignore_index=True )# sheet_df
    return s_df






def getrowrange(wb,s):
    '''
    Args: 
    wb - workbook obj of excel
    s - sheet of the input excel
    Returns:
    dictionary of entity and value as list with 2 items. 1st item is row where 'Pos' occurs. 2nd item is row where 'Go to Top' occurs.
    '''

    sheet_obj =  wb[s]
    maximum_row = sheet_obj.max_row

    ent_dic = {}

    for i in range(1, maximum_row+1):

        cell_obj = sheet_obj.cell(i,pos_col)
        if cell_obj.value == 'Pos':
            entity = sheet_obj.cell(i-3,2) # its 3 rows above Pos word
            row_coord = []
            pos_row=cell_obj.row
            row_coord.append(pos_row)
            j=i
            cell_obj_findtop = sheet_obj.cell(j,pos_col)
            while cell_obj_findtop.value != 'Go to top':
                j += 1
                cell_obj_findtop = sheet_obj.cell(j,pos_col)
            top_row = cell_obj_findtop.row
            row_coord.append(top_row-1) # -1 is for the row before 'Go to top'
            
            ent_dic[entity.value] = row_coord
    
    return ent_dic


def get_pos_sheets(wb, txt):
    '''
    Return all sheets with 'Pos' in specified column 
    Arg:
    wb - workbook to look into
    txt - string to look for
    global variable - pos_col - is used here.
    Returns:
    list of worksheets with word Pos in pos_col column
    '''

    all_sheets = wb.sheetnames

    pos_sheets = []

    for sheet in all_sheets:
        sheet_obj = wb[sheet]
        maxi_row = sheet_obj.max_row

        for i in range(1, maxi_row+1):
            cell_obj = sheet_obj.cell(i,pos_col)
            if cell_obj.value == txt:
                print('present') # logger would be good
                pos_sheets.append(sheet)
                break
    
    return pos_sheets










if __name__ == '__main__': main()
